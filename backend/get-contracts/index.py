"""
Договоры (Admin Panel).
GET  /?pwd=X                         — список всех
GET  /?pwd=X&id=N                    — конкретный
GET  /?pwd=X&report=1&month=YYYY-MM  — Excel отчёт (только оплаченные)
GET  /?pwd=X&staff_id=N              — список по сотруднику
PUT  /?pwd=X&id=N                    — изменить статус / paid / expenses
DELETE /?pwd=X&id=N&confirm=yes      — удалить подписанный договор
"""
import base64
import io
import json
import os
from datetime import datetime, timezone
import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "GET, PUT, DELETE, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type",
    "Content-Type": "application/json",
}


def db():
    return psycopg2.connect(os.environ["DATABASE_URL"])


def s():
    return os.environ.get("MAIN_DB_SCHEMA", "public")


def auth(pwd: str) -> bool:
    return pwd.lower() == os.environ.get("ADMIN_PASSWORD", "Qwert12345").lower()


def make_excel(rows: list, month: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    gold = "FFB300"
    dark = "1A1A1A"
    mid  = "2A2A2A"
    white = "FFFFFF"
    green_fill = PatternFill("solid", fgColor="1A3A2A")
    red_fill   = PatternFill("solid", fgColor="3A1A1A")

    # Заголовок
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"Финансовый отчёт Global Renta — {month or 'все периоды'}"
    title_cell.font = Font(bold=True, size=14, color=gold)
    title_cell.fill = PatternFill("solid", fgColor=dark)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Шапка таблицы
    headers = ["№", "Клиент", "Телефон", "Email", "Сотрудник",
               "КП / Событие", "Дата подписания", "Дата оплаты",
               "Оплата", "Сумма, ₽", "Расходы, ₽", "Прибыль, ₽"]
    header_fill = PatternFill("solid", fgColor=gold)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color=dark, size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24

    thin = Side(style="thin", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [6, 24, 16, 24, 18, 26, 18, 18, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total_rev = total_exp = total_profit = 0

    for idx, r in enumerate(rows):
        row_num = idx + 3
        cid, full_name, company_name, client_type, phone, email, signed_at, paid_at, \
            title, total, event_date, expenses_json, expenses_total, payment_method, staff_name = r

        client = full_name if client_type == "individual" else company_name
        total_val = float(total or 0)
        exp_val   = float(expenses_total or 0)
        profit    = total_val - exp_val
        pay_label = "Безнал" if payment_method == "invoice" else "Наличные"

        total_rev    += total_val
        total_exp    += exp_val
        total_profit += profit

        def fmt_dt(v):
            if not v: return ""
            try: return str(v)[:16].replace("T", " ")
            except: return str(v)[:16]

        row_fill = PatternFill("solid", fgColor="1C1C1C" if idx % 2 == 0 else "222222")
        vals = [
            cid, client, phone or "", email or "", staff_name or "Администратор",
            title or "", fmt_dt(signed_at), fmt_dt(paid_at),
            pay_label, total_val, exp_val, profit
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(color=white, size=9)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if col in (10, 11, 12):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col == 12:
                cell.fill = green_fill if profit >= 0 else red_fill
                cell.font = Font(color="00FF88" if profit >= 0 else "FF6666", bold=True, size=9)

    # Итоговая строка
    tot_row = len(rows) + 3
    ws.merge_cells(f"A{tot_row}:I{tot_row}")
    tot_label = ws[f"A{tot_row}"]
    tot_label.value = "ИТОГО"
    tot_label.font = Font(bold=True, color=gold, size=11)
    tot_label.fill = PatternFill("solid", fgColor=dark)
    tot_label.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[tot_row].height = 22

    for col, val in [(10, total_rev), (11, total_exp), (12, total_profit)]:
        cell = ws.cell(row=tot_row, column=col, value=val)
        cell.font = Font(bold=True, color=gold if col != 12 else ("00FF88" if total_profit >= 0 else "FF6666"), size=11)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = border

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def handler(event: dict, context) -> dict:
    """Управление договорами: список, детали, оплата, расходы, удаление, Excel-отчёт."""
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}

    qp = event.get("queryStringParameters") or {}
    pwd = qp.get("pwd", "")
    if not auth(pwd):
        return {"statusCode": 401, "headers": CORS, "body": json.dumps({"error": "Unauthorized"})}

    method = event.get("httpMethod", "GET")
    schema = s()
    conn = db()
    cur = conn.cursor()

    try:
        # ── EXCEL ОТЧЁТ ────────────────────────────────────────────────
        if method == "GET" and qp.get("report") == "1":
            month = qp.get("month", "")
            staff_id = qp.get("staff_id", "")
            conditions = ["c.paid = true", "c.signed_at IS NOT NULL"]
            params = []
            if month:
                conditions.append("to_char(c.paid_at, 'YYYY-MM') = %s")
                params.append(month)
            if staff_id:
                conditions.append("c.staff_id = %s")
                params.append(int(staff_id))
            where = " AND ".join(conditions)

            cur.execute(
                f"""SELECT c.id, c.full_name, c.company_name, c.client_type,
                    c.phone, c.email, c.signed_at, c.paid_at,
                    q.title, q.total, q.event_date,
                    c.expenses, c.expenses_total, c.payment_method,
                    s.name as staff_name
                FROM {schema}.contracts c
                JOIN {schema}.quotes q ON q.id = c.quote_id
                LEFT JOIN {schema}.staff s ON s.id = c.staff_id
                WHERE {where}
                ORDER BY c.paid_at DESC""",
                params
            )
            rows = cur.fetchall()
            xlsx_bytes = make_excel(rows, month)
            b64 = base64.b64encode(xlsx_bytes).decode()
            fname = f"report_{month or 'all'}.xlsx"
            return {
                "statusCode": 200,
                "headers": {
                    **CORS,
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Disposition": f'attachment; filename="{fname}"',
                    "Content-Transfer-Encoding": "base64",
                },
                "body": b64,
                "isBase64Encoded": True,
            }

        # ── GET список / детали ────────────────────────────────────────
        if method == "GET":
            cid = qp.get("id")
            staff_id = qp.get("staff_id", "")
            if cid:
                cur.execute(
                    f"""SELECT c.id, c.quote_id, c.client_type,
                        c.full_name, c.passport_series, c.passport_number, c.passport_issued, c.passport_date,
                        c.birth_date, c.address,
                        c.company_name, c.inn, c.kpp, c.ogrn, c.legal_address, c.director,
                        c.phone, c.email, c.passport_file_url, c.status, c.created_at,
                        q.title, q.items, q.days, q.delivery, q.delivery_price, q.extras, q.total,
                        q.event_date, q.delivery_address
                    FROM {schema}.contracts c
                    JOIN {schema}.quotes q ON q.id = c.quote_id
                    WHERE c.id = %s""", (int(cid),)
                )
                row = cur.fetchone()
                if not row:
                    return {"statusCode": 404, "headers": CORS, "body": json.dumps({"error": "Not found"})}
                keys = ["id","quote_id","client_type","full_name","passport_series","passport_number",
                        "passport_issued","passport_date","birth_date","address",
                        "company_name","inn","kpp","ogrn","legal_address","director",
                        "phone","email","passport_file_url","status","created_at",
                        "quote_title","items","days","delivery","delivery_price","extras","total",
                        "event_date","delivery_address"]
                result = dict(zip(keys, row))
                result["created_at"] = str(result["created_at"])
                return {"statusCode": 200, "headers": CORS, "body": json.dumps(result, default=str)}
            else:
                extra_where = ""
                params = []
                if staff_id:
                    extra_where = "WHERE c.staff_id = %s"
                    params.append(int(staff_id))
                cur.execute(
                    f"""SELECT c.id, c.quote_id, c.client_type, c.full_name, c.company_name,
                        c.phone, c.email, c.status, c.created_at,
                        q.title, q.total, c.passport_file_url,
                        q.event_date, q.delivery_address,
                        c.signed_at, c.contract_pdf_url,
                        c.payment_method, c.invoice_pdf_url, c.invoice_total,
                        c.paid, c.paid_at, c.expenses, c.expenses_total,
                        c.staff_id, s.name as staff_name
                    FROM {schema}.contracts c
                    JOIN {schema}.quotes q ON q.id = c.quote_id
                    LEFT JOIN {schema}.staff s ON s.id = c.staff_id
                    {extra_where}
                    ORDER BY c.created_at DESC""",
                    params
                )
                keys = ["id","quote_id","client_type","full_name","company_name",
                        "phone","email","status","created_at","quote_title","total","passport_file_url",
                        "event_date","delivery_address","signed_at","contract_pdf_url",
                        "payment_method","invoice_pdf_url","invoice_total",
                        "paid","paid_at","expenses","expenses_total","staff_id","staff_name"]
                rows = [dict(zip(keys, r)) for r in cur.fetchall()]
                for r in rows:
                    r["created_at"] = str(r["created_at"])
                    r["signed_at"]  = str(r["signed_at"]) if r["signed_at"] else None
                    r["paid_at"]    = str(r["paid_at"])   if r["paid_at"]   else None
                    r["expenses"]   = r["expenses"] if isinstance(r["expenses"], list) else []
                    r["expenses_total"] = float(r["expenses_total"] or 0)
                return {"statusCode": 200, "headers": CORS, "body": json.dumps(rows, default=str)}

        # ── PUT: статус / paid / расходы ───────────────────────────────
        if method == "PUT":
            cid = int(qp.get("id", 0))
            body = json.loads(event.get("body") or "{}")

            if "paid" in body:
                paid_val = bool(body["paid"])
                paid_at  = datetime.now(timezone.utc) if paid_val else None
                # При оплате проставляем staff_id из КП если ещё не стоит
                cur.execute(
                    f"UPDATE {schema}.contracts SET paid=%s, paid_at=%s WHERE id=%s",
                    (paid_val, paid_at, cid)
                )
                # Подтягиваем staff_id из quotes если нет в contracts
                cur.execute(
                    f"UPDATE {schema}.contracts c SET staff_id = q.staff_id "
                    f"FROM {schema}.quotes q WHERE q.id = c.quote_id AND c.id = %s AND c.staff_id IS NULL",
                    (cid,)
                )
                conn.commit()
                return {"statusCode": 200, "headers": CORS, "body": json.dumps({"ok": True})}

            if "expenses" in body:
                expenses = body["expenses"]
                total_exp = sum(float(e.get("amount", 0)) for e in expenses)
                cur.execute(
                    f"UPDATE {schema}.contracts SET expenses=%s, expenses_total=%s WHERE id=%s",
                    (json.dumps(expenses, ensure_ascii=False), total_exp, cid)
                )
                conn.commit()
                return {"statusCode": 200, "headers": CORS, "body": json.dumps({"ok": True, "expenses_total": total_exp})}

            new_status = body.get("status", "reviewed")
            cur.execute(f"UPDATE {schema}.contracts SET status=%s WHERE id=%s", (new_status, cid))
            conn.commit()
            return {"statusCode": 200, "headers": CORS, "body": json.dumps({"ok": True})}

        # ── DELETE ─────────────────────────────────────────────────────
        if method == "DELETE":
            cid = int(qp.get("id", 0))
            if qp.get("confirm") != "yes":
                return {"statusCode": 400, "headers": CORS, "body": json.dumps({"error": "Передайте confirm=yes"})}
            cur.execute(f"SELECT signed_at FROM {schema}.contracts WHERE id=%s", (cid,))
            row = cur.fetchone()
            if not row:
                return {"statusCode": 404, "headers": CORS, "body": json.dumps({"error": "Не найден"})}
            if not row[0]:
                return {"statusCode": 400, "headers": CORS, "body": json.dumps({"error": "Только подписанные договоры можно удалять"})}
            cur.execute(f"UPDATE {schema}.contracts SET status='deleted' WHERE id=%s", (cid,))
            conn.commit()
            return {"statusCode": 200, "headers": CORS, "body": json.dumps({"ok": True})}

        return {"statusCode": 405, "headers": CORS, "body": json.dumps({"error": "Method not allowed"})}

    finally:
        cur.close()
        conn.close()
