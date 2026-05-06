"""
Экспорт каталога оборудования в Excel (.xlsx).
GET /?pwd=X&mode=basic — базовый (название, категория, цена)
GET /?pwd=X&mode=full  — расширенный (+ характеристики, теги, описание, рейтинг)
"""
import json, os
from datetime import datetime
from itertools import groupby
import psycopg2
import boto3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "GET, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type",
    "Content-Type": "application/json",
}


def get_db():
    return psycopg2.connect(os.environ["DATABASE_URL"])


def get_s3():
    return boto3.client("s3", endpoint_url="https://bucket.poehali.dev",
                        aws_access_key_id=os.environ["AWS_ACCESS_KEY_ID"],
                        aws_secret_access_key=os.environ["AWS_SECRET_ACCESS_KEY"])


def style_cell(cell, font, fill, alignment, border):
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border


def build_excel_basic(rows: list) -> bytes:
    """Базовый: №, Название, Категория, Подкатегория, Цена, Единица, Популярное"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каталог"

    header_fill = PatternFill("solid", fgColor="1E293B")
    cat_fill    = PatternFill("solid", fgColor="F1F5F9")
    even_fill   = PatternFill("solid", fgColor="FFFFFF")
    odd_fill    = PatternFill("solid", fgColor="F8FAFC")
    title_fill  = PatternFill("solid", fgColor="FEF3C7")

    hfont  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    cfont  = Font(bold=True, color="1E293B", size=9,  name="Calibri")
    bfont  = Font(color="334155",  size=9, name="Calibri")
    pfont  = Font(bold=True, color="166534", size=9, name="Calibri")
    tfont  = Font(bold=True, color="1E293B", size=13, name="Calibri")

    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    c_al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_al   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    COLS = 7
    ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
    tc = ws["A1"]
    tc.value = f"Каталог оборудования Stage Sound — {datetime.now().strftime('%d.%m.%Y')}  |  Базовый"
    tc.font = tfont; tc.fill = title_fill; tc.alignment = c_al
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{get_column_letter(COLS)}2")
    ws.row_dimensions[2].height = 6

    headers    = ["№", "Название", "Категория", "Подкатегория", "Цена, руб./день", "Ед.", "Хит"]
    col_widths = [5,   42,          22,           22,             18,                8,     6]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = hfont; cell.fill = header_fill
        cell.alignment = c_al; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 18

    rows_sorted = sorted(rows, key=lambda r: (r["category"] or "", r["subcategory"] or "", r["sort_order"] or 0))
    dr = 4; num = 1

    for cat_name, cat_items in groupby(rows_sorted, key=lambda r: r["category"] or "Без категории"):
        items = list(cat_items)
        ws.merge_cells(f"A{dr}:{get_column_letter(COLS)}{dr}")
        cc = ws.cell(row=dr, column=1, value=f"  {cat_name.upper()}  ({len(items)} позиций)")
        cc.font = cfont; cc.fill = cat_fill; cc.alignment = l_al; cc.border = border
        ws.row_dimensions[dr].height = 15
        dr += 1

        for item in items:
            fill = even_fill if num % 2 == 0 else odd_fill
            vals = [
                num,
                item.get("name") or "",
                item.get("category") or "",
                item.get("subcategory") or "",
                item.get("price") or 0,
                item.get("unit") or "день",
                "★" if item.get("popular") else "",
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=dr, column=ci, value=v)
                cell.fill = fill; cell.border = border
                if ci == 5:
                    cell.font = pfont; cell.alignment = c_al
                elif ci in (1, 6, 7):
                    cell.font = bfont; cell.alignment = c_al
                else:
                    cell.font = bfont; cell.alignment = l_al
            ws.row_dimensions[dr].height = 14
            dr += 1; num += 1

    ws.freeze_panes = "A4"
    ws.merge_cells(f"A{dr}:{get_column_letter(COLS)}{dr}")
    fc = ws.cell(row=dr, column=1, value=f"Итого: {num - 1} позиций")
    fc.font = Font(bold=True, size=9, name="Calibri"); fc.alignment = l_al

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def build_excel_full(rows: list) -> bytes:
    """Расширенный: + описание, характеристики, теги, рейтинг, отзывы"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каталог (расширенный)"

    header_fill = PatternFill("solid", fgColor="1E293B")
    cat_fill    = PatternFill("solid", fgColor="EFF6FF")
    even_fill   = PatternFill("solid", fgColor="FFFFFF")
    odd_fill    = PatternFill("solid", fgColor="F8FAFC")
    title_fill  = PatternFill("solid", fgColor="FEF3C7")
    spec_fill   = PatternFill("solid", fgColor="F0FDF4")

    hfont = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    cfont = Font(bold=True, color="1E3A8A", size=9,  name="Calibri")
    bfont = Font(color="334155",  size=9,  name="Calibri")
    pfont = Font(bold=True, color="166534", size=9,  name="Calibri")
    sfont = Font(color="4B5563", size=8,  name="Calibri", italic=True)
    tfont = Font(bold=True, color="1E293B", size=13, name="Calibri")

    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    c_al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_al   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    COLS = 11
    ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
    tc = ws["A1"]
    tc.value = f"Каталог оборудования Stage Sound — {datetime.now().strftime('%d.%m.%Y')}  |  Расширенный"
    tc.font = tfont; tc.fill = title_fill; tc.alignment = c_al
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{get_column_letter(COLS)}2")
    ws.row_dimensions[2].height = 6

    headers    = ["№", "Название", "Категория", "Подкатегория", "Цена, руб./день", "Ед.",
                  "Хит", "Рейтинг", "Отзывы", "Теги", "Описание"]
    col_widths = [5,   42,          22,           22,             18,                8,
                  6,   9,           9,            28,             55]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = hfont; cell.fill = header_fill
        cell.alignment = c_al; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 18

    rows_sorted = sorted(rows, key=lambda r: (r["category"] or "", r["subcategory"] or "", r["sort_order"] or 0))
    dr = 4; num = 1

    for cat_name, cat_items in groupby(rows_sorted, key=lambda r: r["category"] or "Без категории"):
        items = list(cat_items)
        ws.merge_cells(f"A{dr}:{get_column_letter(COLS)}{dr}")
        cc = ws.cell(row=dr, column=1, value=f"  {cat_name.upper()}  ({len(items)} позиций)")
        cc.font = cfont; cc.fill = cat_fill; cc.alignment = l_al; cc.border = border
        ws.row_dimensions[dr].height = 15
        dr += 1

        for item in items:
            fill = even_fill if num % 2 == 0 else odd_fill

            # Теги
            tags = item.get("tags") or []
            if isinstance(tags, str):
                import json as _json
                try: tags = _json.loads(tags)
                except: tags = []
            tags_str = ", ".join(tags) if tags else ""

            vals = [
                num,
                item.get("name") or "",
                item.get("category") or "",
                item.get("subcategory") or "",
                item.get("price") or 0,
                item.get("unit") or "день",
                "★" if item.get("popular") else "",
                round(float(item.get("rating") or 0), 1) if item.get("rating") else "",
                item.get("reviews") or "",
                tags_str,
                item.get("description") or "",
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=dr, column=ci, value=v)
                cell.fill = fill; cell.border = border
                if ci == 5:
                    cell.font = pfont; cell.alignment = c_al
                elif ci in (1, 6, 7, 8, 9):
                    cell.font = bfont; cell.alignment = c_al
                elif ci == 10:
                    cell.font = sfont; cell.alignment = l_al
                else:
                    cell.font = bfont; cell.alignment = l_al
            ws.row_dimensions[dr].height = 14
            dr += 1

            # Характеристики (specs) — дополнительные строки под позицией
            specs = item.get("specs") or {}
            if isinstance(specs, str):
                import json as _json
                try: specs = _json.loads(specs)
                except: specs = {}
            if specs and isinstance(specs, dict):
                for key, val in specs.items():
                    ws.merge_cells(f"B{dr}:B{dr}")
                    kc = ws.cell(row=dr, column=2, value=f"    › {key}: {val}")
                    kc.font = sfont; kc.fill = spec_fill
                    kc.alignment = l_al; kc.border = border
                    for ci in [1] + list(range(3, COLS + 1)):
                        cell = ws.cell(row=dr, column=ci)
                        cell.fill = spec_fill; cell.border = border
                    ws.row_dimensions[dr].height = 12
                    dr += 1

            num += 1

    ws.freeze_panes = "A4"
    ws.merge_cells(f"A{dr}:{get_column_letter(COLS)}{dr}")
    fc = ws.cell(row=dr, column=1, value=f"Итого: {num - 1} позиций")
    fc.font = Font(bold=True, size=9, name="Calibri"); fc.alignment = l_al

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def handler(event: dict, context) -> dict:
    """Генерирует Excel каталога. mode=basic — базовый, mode=full — расширенный с характеристиками."""
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}

    qp   = event.get("queryStringParameters") or {}
    pwd  = qp.get("pwd", "")
    mode = qp.get("mode", "basic")

    if pwd.lower() != os.environ.get("ADMIN_PASSWORD", "").lower():
        return {"statusCode": 401, "headers": CORS, "body": json.dumps({"error": "Unauthorized"})}

    schema = os.environ.get("MAIN_DB_SCHEMA", "public")
    conn = get_db(); cur = conn.cursor()
    cur.execute(f"""
        SELECT id, name, category, subcategory, price, unit, popular, description,
               tags, specs, rating, reviews, is_active, sort_order
        FROM {schema}.equipment
        WHERE is_active = TRUE
        ORDER BY category, subcategory, sort_order, name
    """)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    cur.close(); conn.close()

    xlsx  = build_excel_full(rows) if mode == "full" else build_excel_basic(rows)
    label = "full" if mode == "full" else "basic"
    fname = f"catalog_{label}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    key   = f"exports/{fname}"

    s3 = get_s3()
    s3.put_object(Bucket="files", Key=key, Body=xlsx,
                  ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    cdn = f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{key}"
    return {"statusCode": 200, "headers": CORS,
            "body": json.dumps({"ok": True, "url": cdn, "count": len(rows)})}
